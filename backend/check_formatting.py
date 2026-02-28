#!/usr/bin/env python3
"""Check Excel formatting"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border

OUTPUT_PATH = "../outputs/test_output.xlsx"

wb = openpyxl.load_workbook(OUTPUT_PATH)
ws = wb.active

print("=== FORMATTING CHECK ===\n")

# Check header formatting
print("--- HEADER ROW (Row 1) ---")
cell_1 = ws.cell(row=1, column=1)  # Ghi chú
cell_2 = ws.cell(row=1, column=9)  # Người đặt hàng (should be red)
cell_3 = ws.cell(row=1, column=14)  # TỈNH (should be red)

print(f"Column 1 (Ghi chú):")
print(f"  Font: {cell_1.font.name}, Size: {cell_1.font.size}, Bold: {cell_1.font.bold}")
print(f"  Fill: {cell_1.fill.start_color.rgb if hasattr(cell_1.fill.start_color, 'rgb') else cell_1.fill.start_color}")
print(f"  Alignment: {cell_1.alignment}")

print(f"\nColumn 9 (Người đặt hàng - should be RED):")
print(f"  Font: {cell_2.font.name}, Size: {cell_2.font.size}, Color: {cell_2.font.color}")

print(f"\nColumn 14 (TỈNH - should be RED):")
print(f"  Font: {cell_3.font.name}, Size: {cell_3.font.size}, Color: {cell_3.font.color}")

# Check data row formatting
print("\n--- DATA ROW (Row 2) ---")
cell_data_1 = ws.cell(row=2, column=2)  # STT (should be centered)
cell_data_2 = ws.cell(row=2, column=14)  # TỈNH (should be red)

print(f"Column 2 (STT - should be CENTERED):")
print(f"  Font: {cell_data_1.font.name}, Size: {cell_data_1.font.size}")
print(f"  Alignment: {cell_data_1.alignment}")

print(f"\nColumn 14 (TỈNH data - should be RED):")
print(f"  Font: {cell_data_2.font.name}, Size: {cell_data_2.font.size}")
print(f"  Color: {cell_data_2.font.color}")

print("\n" + "="*50)
