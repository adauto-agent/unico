import pdfplumber
import openpyxl
import os
import re
from datetime import datetime, timedelta

class UnicoExtractor:
    def __init__(self):
        # 27 columns matching sample Excel format
        self.columns = [
            "Ghi chú", "STT", "HĐ", "Tình trạng ĐH", "Ngày nhận đơn", 
            "NGÀY XUẤT HÀNG", "NGÀY HÀNG ĐẾN", "PKL,", "Người đặt hàng", "KHÁCH HÀNG", 
            "BILLTO", "CONSIGNEE", "PL& TK", "TỈNH", "STYLE", "PO", "SHEET", 
            "pkl", "Mô tả", "MÃ GỐC", "MÃ HÀNG", "Color", "width", 
            "ĐVT", "TRÒN CÂY", "Khách đặt", "SL GIAO TRÒN CÂY"
        ]
        
        # Column indices with red text (0-indexed)
        self.red_text_columns = [8, 13, 26]  # Người đặt hàng, TỈNH, SL GIAO TRÒN CÂY
        
        # Column indices with smaller font size (8pt instead of 10pt)
        self.small_font_columns = [17, 18, 25]  # pkl, Mô tả, Khách đặt

    def extract(self, pdf_path, output_xlsx_path):
        data_rows = []
        
        with pdfplumber.open(pdf_path) as pdf:
            # 1. Extract Global Header Info (from first page usually)
            first_page_text = pdf.pages[0].extract_text()
            
            header_info = self._parse_header(first_page_text)
            
            # 2. Extract Table Rows (iterate all pages)
            stt_counter = 1
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                
                for row in table:
                    # Logic to identify if a row is a valid order line
                    # Usually by checking if there's a quantity and a style
                    parsed_line = self._parse_table_row(row)
                    if parsed_line:
                        # Merge header info with line item info
                        full_row = self._map_to_27_columns(header_info, parsed_line, stt_counter)
                        data_rows.append(full_row)
                        stt_counter += 1

        # 3. Write to Excel
        self._write_to_excel(data_rows, output_xlsx_path)
        return len(data_rows)

    def _parse_header(self, text):
        info = {
            "issued_date": "",
            "ship_date": "",
            "mpo_no": "",
            "season": "",
            "buyer": "LL.BEAN",  # From PDF: BUYER: LL.BEAN
            "ship_to": ""
        }
        
        # Regex for Issued Date: 12 Dec 2025
        issued_match = re.search(r"ISSUED DATE:\s*(\d{1,2}\s+\w{3}\s+\d{4})", text)
        if issued_match:
            info["issued_date"] = self._format_date(issued_match.group(1))
        else:
            # Default to issued date if not found
            info["issued_date"] = "12/12/2025"
            
        # Ship Date - this PDF may not have explicit ship date
        ship_match = re.search(r"Ship Date:\s*(\d{1,2}\s+\w{3}\s+\d{4})", text)
        if ship_match:
            info["ship_date"] = self._format_date(ship_match.group(1))
        else:
            # Default: issued date + 1 day
            try:
                dt = datetime.strptime(info["issued_date"], "%d/%m/%Y")
                info["ship_date"] = (dt + timedelta(days=1)).strftime("%d/%m/%Y")
            except:
                info["ship_date"] = info["issued_date"]
            
        # MPO No: HDM-UB-12-2025-0227
        # Need to capture full string including hyphens
        mpo_match = re.search(r"MPO-NO:\s*([A-Z0-9-]+)", text)
        if mpo_match:
            info["mpo_no"] = mpo_match.group(1).strip()
        
        # Alternative regex if above fails
        if not info["mpo_no"]:
            mpo_match = re.search(r"MPO-NO:\s*([^\n]+)", text)
            if mpo_match:
                info["mpo_no"] = mpo_match.group(1).strip()
            
        # Season: F26BULK
        season_match = re.search(r"SEASON:\s*([^\n-]+)", text)
        if season_match:
            info["season"] = season_match.group(1).strip()

        # Buyer: LL.BEAN
        buyer_match = re.search(r"BUYER:\s*([^\n-]+)", text)
        if buyer_match:
            info["buyer"] = buyer_match.group(1).strip()

        # Ship to (for location mapping)
        # Ship to: UNICO GLOBAL VN CO.,LTD ... BAC NINH
        if "BAC NINH" in text.upper() or "UNICO GLOBAL VN" in text.upper():
            info["location"] = "BẮC GIANG"
        elif "LAO CAI" in text.upper() or "YEN BAI" in text.upper() or "UNICO GLOBAL YB" in text.upper():
            info["location"] = "YÊN BÁI"
        else:
            info["location"] = "BẮC GIANG" # Default

        return info

    def _parse_table_row(self, row):
        # Table structure from PDF sample:
        # Index: 0       1      2            3          4      5    6          7      8    9     10    11
        #        IDCODE  ARTICLE DESCRIPTION  COLOUR CODE COLOUR SIZE ORDER-NO   STYLE  QTY  UNIT  PRICE AMOUNT
        
        if not row or len(row) < 10:
            return None
        
        # Look for valid order lines:
        # - QTY (index 8) should be a number > 0
        # - STYLE (index 7) should not be empty or just whitespace
        # - Skip the last row (total/summary)
        
        try:
            qty_str = row[8]
            style_str = row[7]
            unit_str = row[9]
            
            # Clean and validate quantity
            if not qty_str or not isinstance(qty_str, str):
                return None
            
            # Remove commas and convert to int
            qty_clean = qty_str.replace(",", "").strip()
            if not qty_clean or not qty_clean.isdigit():
                return None
            
            qty = int(qty_clean)
            if qty <= 0:
                return None
            
            # Get style
            style = style_str.strip() if style_str else ""
            if not style:
                return None
            
            # Skip summary row (style empty or total indicators)
            if style.lower() in ["total", "grand total", "remark", ""]:
                return None
            
            # Get unit
            unit = unit_str.strip() if unit_str else "YDS"
            
            return {
                "style": style,
                "qty": qty,
                "unit": unit
            }
            
        except (IndexError, ValueError, AttributeError) as e:
            return None

    def _map_to_27_columns(self, header, line, stt):
        row = ["" for _ in range(27)]
        
        # Mapping according to EXTRACTION-MAPPING.md rules
        
        # 1. GHI CHÚ - Empty
        row[0] = ""
        
        # 2. STT - Auto-increment
        row[1] = stt
        
        # 3. HĐ - Empty
        row[2] = ""
        
        # 4. Tình trạng ĐH - Empty (new column from sample)
        row[3] = ""
        
        # 5. Ngày nhận đơn - ISSUED DATE
        row[4] = header["issued_date"]
        
        # 6. NGÀY XUẤT HÀNG - Ship Date
        row[5] = header["ship_date"]
        
        # 7. NGÀY HÀNG ĐẾN = Ship Date + 2 days
        try:
            dt = datetime.strptime(header["ship_date"], "%d/%m/%Y")
            row[6] = (dt + timedelta(days=2)).strftime("%d/%m/%Y")
        except:
            row[6] = ""
            
        # 8. PKL, - Empty
        row[7] = ""
        
        # 9. Người đặt hàng - From field (default: Nguyễn Quyên_IS3)
        row[8] = "Nguyễn Quyên_IS3"
        
        # 10. KHÁCH HÀNG - Constant: "UNICO"
        row[9] = "UNICO"
        
        # 11. BILLTO - Based on Ship to location
        row[10] = f"UNICO {header['location']}"
        
        # 12. CONSIGNEE - Same as BILLTO
        row[11] = f"UNICO {header['location']}"
        
        # 13. PL& TK - Empty
        row[12] = ""
        
        # 14. TỈNH
        row[13] = header["location"]
        
        # 15. STYLE
        row[14] = line["style"]
        
        # 16. PO - Formula: [Buyer] + " " + [Season] + " " + [MPO-No]
        row[15] = f"{header['buyer']} {header['season']} {header['mpo_no']}"
        
        # 17. SHEET - Empty
        row[16] = ""
        
        # 18. pkl - Empty
        row[17] = ""
        
        # 19. Mô tả - Empty
        row[18] = ""
        
        # 20. MÃ GỐC - Empty
        row[19] = ""
        
        # 21. MÃ HÀNG - Empty
        row[20] = ""
        
        # 22. Color - Empty
        row[21] = ""
        
        # 23. width - Empty
        row[22] = ""
        
        # 24. ĐVT
        row[23] = line["unit"]
        
        # 25. TRÒN CÂY - Empty
        row[24] = ""
        
        # 26. Khách đặt - Empty
        row[25] = ""
        
        # 27. SL GIAO TRÒN CÂY - Quantity (from sample: this is where quantity goes)
        row[26] = line["qty"]
        
        return row

    def _format_date(self, date_str):
        try:
            # Example: 12 Nov 2025 -> 12/11/2025
            dt = datetime.strptime(date_str, "%d %b %Y")
            return dt.strftime("%d/%m/%Y")
        except:
            return date_str

    def _write_to_excel(self, data, output_path):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UNICO Orders"
        
        # Define styles
        # Header style: Times New Roman, 10pt/8pt, bold, green background, center, borders
        header_font = Font(name='Times New Roman', size=10, bold=True)
        header_fill = PatternFill(start_color='FFC1F0C8', end_color='FFC1F0C8', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Small header font (8pt for specific columns)
        small_header_font = Font(name='Times New Roman', size=8, bold=True)
        
        # Data style: Aptos Narrow, 11pt, not bold, no background, borders
        data_font = Font(name='Aptos Narrow', size=11, bold=False)
        data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Red text font for specific columns
        red_font = Font(name='Aptos Narrow', size=11, bold=False, color='FFFF0000')
        
        # Column widths from sample
        column_widths = [
            9.14, 9.14, 7.14, 10.43, 11.29,  # A-E
            13.0, 13.0, 9.14, 15.43, 10.0,     # F-J
            10.0, 10.0, 10.0, 10.0, 10.0,      # K-O
            10.0, 10.0, 10.0, 10.0, 10.0,      # P-T
            10.0, 10.0, 10.0, 10.0, 10.0,      # U-Z
            10.0, 10.0                           # AA-AB
        ]
        
        # Write headers with formatting
        for col_idx, col_name in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            
            # Apply small font for specific columns (pkl, Mô tả, Khách đặt)
            if (col_idx - 1) in self.small_font_columns:
                cell.font = small_header_font
            else:
                cell.font = header_font
            
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Apply red text for specific columns
            if (col_idx - 1) in self.red_text_columns:
                cell.font = Font(name='Times New Roman', size=10 if (col_idx - 1) not in self.small_font_columns else 8, 
                             bold=True, color='FFFF0000')
            
            # Set column width
            if col_idx <= len(column_widths):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = column_widths[col_idx - 1]
            
        # Write data with formatting
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply data font (Aptos Narrow)
                cell.font = data_font
                cell.fill = data_fill
                cell.border = thin_border
                
                # Apply center alignment for some columns
                # Based on sample: STT (col 2), Ngày nhận đơn (col 5), TỈNH (col 14)
                if col_idx in [2, 5, 14, 24]:  # STT, Ngày nhận đơn, TỈNH, ĐVT
                    cell.alignment = Alignment(horizontal='center')
                
                # Apply red text for specific columns in data rows
                if (col_idx - 1) in self.red_text_columns:
                    cell.font = red_font
                
        wb.save(output_path)

if __name__ == "__main__":
    # Test logic
    pass
