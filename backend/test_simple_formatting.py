#!/usr/bin/env python3
"""Test simple formatting"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active

# Test header formatting
header_font = Font(name='Times New Roman', size=10, bold=True, color='FFFF0000')
header_fill = PatternFill(start_color='FFC1F0C8', end_color='FFC1F0C8', fill_type='solid')

cell = ws.cell(row=1, column=1, value="TEST")
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save("/tmp/test_format.xlsx")

# Verify
wb2 = openpyxl.load_workbook("/tmp/test_format.xlsx")
ws2 = wb2.active
cell2 = ws2.cell(row=1, column=1)

print("Cell font:", cell2.font.name, cell2.font.size, cell2.font.bold)
print("Cell fill:", cell2.fill.start_color)
print("Cell alignment:", cell2.alignment)
