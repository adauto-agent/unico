#!/usr/bin/env python3
"""Verify Excel output content"""

import openpyxl

OUTPUT_PATH = "../outputs/test_output.xlsx"

wb = openpyxl.load_workbook(OUTPUT_PATH)
ws = wb.active

print("=== EXCEL OUTPUT VERIFICATION ===\n")
print(f"Sheet name: {ws.title}")
print(f"Total rows (including header): {ws.max_row}")
print(f"Total columns: {ws.max_column}\n")

print("--- HEADERS ---")
headers = [cell.value for cell in ws[1]]
for i, h in enumerate(headers, 1):
    print(f"{i:2}. {h}")

print("\n--- DATA ROWS ---")
for row_idx in range(2, min(6, ws.max_row + 1)):  # First 5 data rows
    row_data = [cell.value for cell in ws[row_idx]]
    print(f"\nRow {row_idx}:")
    print(f"  STT: {row_data[1]}")
    print(f"  NGÀY NHẬN ĐƠN: {row_data[3]}")
    print(f"  NGÀY XUẤT HÀNG: {row_data[4]}")
    print(f"  NGÀY HÀNG ĐẾN: {row_data[5]}")
    print(f"  STYLE: {row_data[13]}")
    print(f"  PO: {row_data[14]}")
    print(f"  SỐ LƯỢNG: {row_data[23]}")

print("\n" + "="*50)
