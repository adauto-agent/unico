#!/usr/bin/env python3
"""Check formatting of sample Excel file"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border

SAMPLE_PATH = "9e17d789-bc2c-42a1-bd97-1169e4f96de6.xlsx"

wb = openpyxl.load_workbook(SAMPLE_PATH)
ws = wb.active

print("=== SAMPLE EXCEL FORMATTING ANALYSIS ===\n")
print(f"Sheet name: {ws.title}")
print(f"Rows: {ws.max_row}, Columns: {ws.max_column}\n")

# Check header row formatting
print("--- HEADER ROW (Row 1) ---")
for col_idx in range(1, min(28, ws.max_column + 1)):
    cell = ws.cell(row=1, column=col_idx)
    print(f"\nColumn {col_idx}: '{cell.value}'")
    
    # Font
    if cell.font:
        print(f"  Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}, Color: {cell.font.color}")
    
    # Fill (background)
    if cell.fill and cell.fill.start_color:
        print(f"  Fill: {cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else cell.fill.start_color}")
    
    # Alignment
    if cell.alignment:
        print(f"  Alignment: H={cell.alignment.horizontal}, V={cell.alignment.vertical}")
    
    # Border
    if cell.border and any([cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]):
        print(f"  Border: Yes")

# Check data row formatting
print("\n\n--- DATA ROW (Row 2) ---")
for col_idx in range(1, min(10, ws.max_column + 1)):
    cell = ws.cell(row=2, column=col_idx)
    print(f"\nColumn {col_idx}: '{cell.value}'")
    
    if cell.font:
        print(f"  Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}")
    
    if cell.fill and cell.fill.start_color:
        print(f"  Fill: {cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else cell.fill.start_color}")
    
    if cell.alignment:
        print(f"  Alignment: H={cell.alignment.horizontal}, V={cell.alignment.vertical}")

# Check column widths
print("\n\n--- COLUMN WIDTHS ---")
for col_idx in range(1, min(10, ws.max_column + 1)):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    width = ws.column_dimensions[col_letter].width
    print(f"Column {col_letter}: {width}")

print("\n" + "="*50)
